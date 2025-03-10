from netmiko import ConnectHandler
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font

# Definir colores para la terminal
AMBAR = "\033[93m"
GREEN = "\033[92m"
RED = "\033[31m"
RESET = "\033[0m"

# Lista de routers
routers = [
    {"device_type": "cisco_ios", "host": "192.168.1.102", "username": "cisco", "password": "cisco", "secret": "cisco"},
    {"device_type": "cisco_ios", "host": "10.0.3.1", "username": "cisco", "password": "cisco", "secret": "cisco"},
    {"device_type": "cisco_ios", "host": "10.0.3.6", "username": "cisco", "password": "cisco", "secret": "cisco"},
]

# Reglas con sus comandos
rules = {
    "HTTPS habilitado": {"command": "show running-config | include ip http secure-server", "expected": "ip http secure-server"},
    "Telnet deshabilitado": {"command": "show running-config | include transport input", "expected": "transport input ssh"},
    "Solo una LoopBack": {"command": "show ip interface brief | include Loopback", "expected": 1},
    "Syslog configurado": {"command": "show running-config | include logging host", "expected": "logging host"},
    "NTP configurado": {"command": "show running-config | include ntp server", "expected": "ntp server"},
    "Redireccion ICMP deshabilitada": {"command": "show running-config | include ip redirects", "expected": "no ip redirects"},
    "Seguridad en puertos Habilitados": {"command": "show running-config | include switchport port-security", "expected": "switchport port-security"},
    "Banner de Advertencia": {"command": "show running-config | include banner login", "expected": "banner login"},
    "Tiempo de inactividad de la consola": {"command": "show running-config | include exec-timeout", "expected": "exec-timeout"},
    "Control de Acceso a la Conosla": {"command": "show running-config | include login local", "expected": "login local"}
}

def get_hostname(connection):
    """Obtiene el hostname del router usando 'show version'"""
    hostname_output = connection.send_command("show version | include uptime")
    hostname_match = re.search(r"(\S+) uptime is", hostname_output)
    return hostname_match.group(1) if hostname_match else "Desconocido"

def audit_routers(routers, rules):
    audit_results = {rule: {} for rule in rules}  # Diccionario organizado por reglas

    for router in routers:
        try:
            print(f"\n{GREEN}[+]{RESET} Conectando a {router['host']}...")
            net_connect = ConnectHandler(**router)
            net_connect.enable()
            hostname = get_hostname(net_connect)
            print(f"{GREEN}[+]{RESET} Iniciando auditoría en el router: {hostname}")

            for rule, details in rules.items():
                output = net_connect.send_command(details["command"])

                # Verificación de cumplimiento (100 = Cumple, 0 = No cumple)
                if "LoopBack" in rule:
                    count = output.count("Loopback")
                    compliance = 100 if count == details["expected"] else 0
                else:
                    compliance = 100 if details["expected"] in output else 0

                print(f"{rule}: {'Cumple' if compliance == 100 else 'No cumple'}")
                audit_results[rule][hostname] = compliance

            net_connect.disconnect()

        except Exception as e:
            print(f"{RED}[+]{RESET} Error conectando al router {router['host']}: {e}")

    # Convertir datos a DataFrame con reglas como FILAS y routers como COLUMNAS
    df = pd.DataFrame(audit_results).T.fillna(0)

    # Agregar fila de calificación individual (promedio de cada router)
    df.loc["Calificación individual"] = df.mean().round(2)

    # Calcular calificación total (promedio de la calificación individual)
    calificacion_total = df.loc["Calificación individual"].mean().round(2)

    # Agregar fila de "Calificación Total" con una sola celda fusionada
    df.loc["Calificación Total"] = [""] * len(df.columns)
    df.at["Calificación Total", "Calificación Total"] = calificacion_total  # Asignar valor en la última celda

    # Guardar en Excel con formato
    excel_filename = "auditoria_demo.xlsx"
    df.to_excel(excel_filename, index=True)

    print(f"\n{AMBAR}[+]{RESET} Auditoría guardada en '{excel_filename}'")

    # Aplicar formato en Excel
    apply_excel_format(excel_filename, len(df.columns))

def apply_excel_format(filename, num_columns):
    """Aplica bordes, alineación y negritas en el archivo de Excel"""
    wb = load_workbook(filename)
    ws = wb.active

    # Definir bordes
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Aplicar bordes y alineación a todas las celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Aplicar negrita a los encabezados de reglas
    for cell in ws["A"]:  
        cell.font = Font(bold=True)

    # Aplicar negrita a la fila de "Calificación Total"
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.font = Font(bold=True)

    # Fusionar todas las columnas en la fila de "Calificación Total"
    ws.merge_cells(start_row=last_row, start_column=2, end_row=last_row, end_column=num_columns)

    # Ajustar el valor de la celda fusionada
    ws.cell(row=last_row, column=2).value = ws.cell(row=last_row, column=num_columns).value
    ws.cell(row=last_row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    # Guardar cambios
    wb.save(filename)
    print(f"{GREEN}[+]{RESET} Formato aplicado en '{filename}'")

audit_routers(routers, rules)
